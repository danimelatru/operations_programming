import random
import pandas as pd
import os
from scheptk.util import write_tag
from scheptk.scheptk import randint
from openpyxl import load_workbook

def cargar_datos_excel(excel_path):
    # Carga y procesa los datos desde un archivo Excel.
    df = pd.read_excel(excel_path, skiprows=2).iloc[:, :10]  # Filtrar columnas relevantes
    df.columns = [
        'Ref_Producto', 'Tiempo_Mecanizado_CNC1', 'Tiempo_Mecanizado_CNC2', 'Tiempo_Mecanizado_CNC3',
        'Tiempo_Soldadura_Puesto1', 'Tiempo_Soldadura_Puesto2', 'Tiempo_Montaje',
        'Tiempo_Preparacion_Soldadura_Puesto1', 'Tiempo_Preparacion_Soldadura_Puesto2', 'Coste_Producto'
    ]
    df = df.dropna()  # Eliminar filas vacías
    df['Tiempo_Soldadura_Puesto1'] += df['Tiempo_Preparacion_Soldadura_Puesto1']
    df['Tiempo_Soldadura_Puesto2'] += df['Tiempo_Preparacion_Soldadura_Puesto2']
    return df

def generar_instancia(nombre_instancia, numero_trabajos):
    # Cargar y procesar los datos desde el Excel
    datos = cargar_datos_excel('Datos_PO.xlsx')
    
    # Asignar las columnas a variables, manteniendo los mismos nombres para compatibilidad
    tm1 = datos['Tiempo_Mecanizado_CNC1'].astype(int).tolist()
    tm2 = datos['Tiempo_Mecanizado_CNC2'].astype(int).tolist()
    tm3 = datos['Tiempo_Mecanizado_CNC3'].astype(int).tolist()

    tsold1 = datos['Tiempo_Soldadura_Puesto1'].astype(int).tolist()
    tsold2 = datos['Tiempo_Soldadura_Puesto2'].astype(int).tolist()

    # Si aún necesitas los tiempos de preparación separados, puedes asignarlos antes de la suma
    ssold1 = datos['Tiempo_Preparacion_Soldadura_Puesto1'].fillna(0).astype(int).tolist()
    ssold2 = datos['Tiempo_Preparacion_Soldadura_Puesto2'].fillna(0).astype(int).tolist()

    tmont = datos['Tiempo_Montaje'].astype(int).tolist()

    pesos = datos['Coste_Producto'].astype(int).tolist()

    # Genera trabajos aleatorios y los guarda en un archivo de texto.
    referencias = random.choices(list(range(1, 200)), k=numero_trabajos)
    datos_instancia = {
        "pt1": [tm1[i-1] for i in referencias],
        "pt2": [tm2[i-1] for i in referencias],
        "pt3": [tm3[i-1] for i in referencias],
        "pt4": [tsold1[i-1] for i in referencias],
        "pt5": [tsold2[i-1] for i in referencias],
        "pt6": [tmont[i-1] for i in referencias],
        "setup1": [ssold1[i-1] for i in referencias],
        "setup2": [ssold2[i-1] for i in referencias],
        "costes": [pesos[i-1] for i in referencias]
    }
    inversos = [1 / c for c in datos_instancia["costes"] if c != 0]
    suma_inversos = sum(inversos)
    w = [round(inv / suma_inversos, 4) for inv in inversos]
    a = calcula_a(datos_instancia)
    d = calcula_d(datos_instancia)

    crear_archivo_txt(nombre_instancia)
    write_tag('JOBS', numero_trabajos, nombre_instancia)
    write_tag('MACHINES', 6, nombre_instancia)

    for key, tag in zip(["pt1", "pt2", "pt3", "pt4", "pt5", "pt6"], 
                         ["PT_MEC1", "PT_MEC2", "PT_MEC3", "PT_SOL1", "PT_SOL2", "PT_MONT"]):
        write_tag(tag, datos_instancia[key], nombre_instancia)

    write_tag('SS_SOL1', datos_instancia["setup1"], nombre_instancia)
    write_tag('SS_SOL2', datos_instancia["setup2"], nombre_instancia)
    write_tag('W', w, nombre_instancia)
    write_tag('A', a, nombre_instancia)
    write_tag('DD', d, nombre_instancia)
    write_tag('REFS', referencias, nombre_instancia)

def calcula_a(datos_instancia):
    # Calcula valores de ocupación inicial "a" para un flowshop híbrido.
    pt_totales = [datos_instancia["pt1"], datos_instancia["pt2"], datos_instancia["pt3"], 
                  datos_instancia["pt4"], datos_instancia["pt5"], datos_instancia["pt6"]]
    a = []
    for i, grupo in enumerate([3, 2, 1]):
        for j in range(grupo):
            sum_pij = sum(pt_totales.pop(0))
            limite_inferior = sum_pij / (6 - i + 1)
            limite_superior = (6 / (6 - i + 1)) * sum_pij
            a.append(int(random.uniform(limite_inferior, limite_superior)))
    return a

def calcula_d(datos_instancia):
    # Calcula fechas de entrega aleatorias.
    pt_totales = [datos_instancia["pt1"], datos_instancia["pt2"], datos_instancia["pt3"], 
                  datos_instancia["pt4"], datos_instancia["pt5"], datos_instancia["pt6"]]
    return [randint(sum(pt[j] for pt in pt_totales), 6 * sum(pt[j] for pt in pt_totales))
            for j in range(len(pt_totales[0]))]

def crear_archivo_txt(nombre):
    #Crea o limpia un archivo de texto.
    if os.path.exists(nombre):
        os.remove(nombre)
    with open(nombre, 'w') as archivo:
        archivo.write(f'# Nombre: {nombre}\n')


def guardar_resultados_excel(excel_path, resultados_despacho, resultados_heuristicas, resultados_metaheuristicas):
    
    from openpyxl import Workbook

    # Si el archivo no existe, crea uno nuevo
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    # Cargar el archivo Excel existente
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Verificar si la hoja "Datos" ya existe, cargándola si es necesario
        wb = load_workbook(excel_path)
        if "Datos" in wb.sheetnames:
            print("Fin.")

        # Guardar los resultados en hojas nuevas o reemplazar las existentes
        pd.DataFrame(resultados_despacho).to_excel(writer, sheet_name='Resultados Reglas Despacho', index=False)
        pd.DataFrame(resultados_heuristicas).to_excel(writer, sheet_name='Resultados Heuristicas', index=False)
        pd.DataFrame(resultados_metaheuristicas).to_excel(writer, sheet_name='Resultados Metaheuristicas', index=False)


